#! python3 mul_table.py for a N*N mulplication table.

import openpyxl

N = int(input("offer an integer( >= 1 ): "))
wb = openpyxl.Workbook()
sheet = wb['Sheet']

for i in range(1, N + 2):
    for j in range(1, N + 2):
        if i == 1 and j == 1:
            pass
        elif i == 1 or j == 1:
            sheet.cell(row=i, column=j).value = max(i, j) - 1
        else:
            sheet.cell(row=i, column=j).value = (i - 1) * (j - 1)

wb.save('mul_table.xlsx')
